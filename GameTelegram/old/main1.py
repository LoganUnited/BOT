import os
import sqlite3
import asyncio
import re
import requests
from datetime import datetime, timedelta, time
from typing import List
from openpyxl import Workbook, load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackContext,
    ConversationHandler,
    MessageHandler,
    filters,
    CallbackQueryHandler
)
import needs  # Импортируем модуль needs.py

# Конфигурация
BOT_TOKEN = "7578255189:AAEkOcP-BmvZnOWXz39iYCop7PgDVMllSMQ"
GET_NICKNAME = 1
DB_VERSION = 5  # Увеличена версия БД
EXPORT_FOLDER = "exports"
MAX_FILE_AGE_DAYS = 31
ALLOWED_LOCATIONS = ["Лес", "Город", "Пещера", "Деревня", "Лагерь", "Храм"]
DEFAULT_LOCATION = "Лагерь"
CHAT_ID_WHITELIST = [123456789, 987654321]
OPENWEATHER_API_KEY = "ВАШ_API_КЛЮЧ_OPENWEATHERMAP"
MOSCOW_TIMEZONE_OFFSET = 3

if not os.path.exists(EXPORT_FOLDER):
    os.makedirs(EXPORT_FOLDER)

# ====================== УПРАВЛЕНИЕ ФАЙЛАМИ =======================
def clean_old_exports():
    now = datetime.now()
    pattern = re.compile(r"players_stats_(\d{4}-\d{2}-\d{2})_\d{2}-\d{2}-\d{2}\.xlsx")
    
    for filename in os.listdir(EXPORT_FOLDER):
        match = pattern.match(filename)
        if match:
            file_date_str = match.group(1)
            try:
                file_date = datetime.strptime(file_date_str, "%Y-%m-%d")
                if (now - file_date) > timedelta(days=MAX_FILE_AGE_DAYS):
                    file_path = os.path.join(EXPORT_FOLDER, filename)
                    os.remove(file_path)
                    print(f"🗑 Удален старый файл: {filename}")
            except Exception as e:
                print(f"Ошибка обработки файла {filename}: {str(e)}")

async def auto_cleanup_task(context: CallbackContext):
    clean_old_exports()
    print("✅ Выполнена очистка старых файлов экспорта")

def get_export_filename() -> str:
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return os.path.join(EXPORT_FOLDER, f"players_stats_{timestamp}.xlsx")

# ====================== МИГРАЦИИ БАЗЫ ДАННЫХ ======================
def get_db_version(conn: sqlite3.Connection) -> int:
    cursor = conn.cursor()
    try:
        cursor.execute("PRAGMA user_version")
        return cursor.fetchone()[0]
    except sqlite3.OperationalError:
        return 0

def set_db_version(conn: sqlite3.Connection, version: int) -> None:
    cursor = conn.cursor()
    cursor.execute(f"PRAGMA user_version = {version}")
    conn.commit()

def migrate_db(conn: sqlite3.Connection) -> None:
    current_version = get_db_version(conn)
    if current_version >= DB_VERSION:
        return

    migrations = [
        '''CREATE TABLE IF NOT EXISTS users (
            player_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER UNIQUE,
            username TEXT UNIQUE,
            score INTEGER DEFAULT 0,
            money INTEGER DEFAULT 0,
            skills INTEGER DEFAULT 1,
            level INTEGER DEFAULT 1,
            experience INTEGER DEFAULT 0,
            location TEXT DEFAULT 'Лагерь',
            last_active TEXT DEFAULT CURRENT_TIMESTAMP,
            is_afk BOOLEAN DEFAULT 0,
            food INTEGER DEFAULT 100,
            water INTEGER DEFAULT 100,
            sleep INTEGER DEFAULT 100,
            last_need_update TEXT DEFAULT CURRENT_TIMESTAMP
        )''',
    ]

    cursor = conn.cursor()
    for i in range(current_version, DB_VERSION):
        if i < len(migrations):
            cursor.executescript(migrations[i])
            set_db_version(conn, i + 1)

def init_db():
    conn = sqlite3.connect('game.db')
    migrate_db(conn)
    conn.close()

# ====================== ЭКСПОРТ/ИМПОРТ В EXCEL ===================
def export_to_excel(filename: str) -> None:
    conn = sqlite3.connect('game.db')
    cursor = conn.cursor()
    cursor.execute('SELECT player_id, username, level, experience, money, skills, location, is_afk FROM users')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Players"
    ws.append(["ID", "Никнейм", "Уровень", "Опыт", "Деньги", "Навыки", "Местоположение", "AFK"])
    
    for row in cursor.fetchall():
        ws.append(row)
    
    wb.save(filename)
    conn.close()

def import_from_excel(filename: str = "players_stats.xlsx") -> None:
    conn = sqlite3.connect('game.db')
    cursor = conn.cursor()
    
    wb = load_workbook(filename)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        player_id, username, level, exp, money, skills, location, is_afk = row
        cursor.execute('''
            UPDATE users SET 
                username = ?, 
                level = ?, 
                experience = ?, 
                money = ?, 
                skills = ?,
                location = ?,
                is_afk = ?
            WHERE player_id = ?
        ''', (username, level, exp, money, skills, location, is_afk, player_id))
    
    conn.commit()
    conn.close()

# ====================== ОСНОВНЫЕ ФУНКЦИИ =========================
async def add_experience(user_id: int, amount: int) -> None:
    with sqlite3.connect('game.db', timeout=10) as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT level, experience FROM users WHERE user_id = ?', (user_id,))
        level, exp = cursor.fetchone()
        
        new_exp = exp + amount
        while True:
            needed = 4 * (level + 1)
            if new_exp < needed: break
            new_exp -= needed
            level += 1
        
        cursor.execute('UPDATE users SET level = ?, experience = ? WHERE user_id = ?', 
                      (level, new_exp, user_id))
        conn.commit()

def user_exists(user_id: int) -> bool:
    with sqlite3.connect('game.db', timeout=10) as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT 1 FROM users WHERE user_id = ?', (user_id,))
        exists = cursor.fetchone() is not None
    return exists

def update_activity(func):
    async def wrapper(update: Update, context: CallbackContext):
        user_id = update.message.from_user.id
        with sqlite3.connect('game.db') as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE users 
                SET last_active=CURRENT_TIMESTAMP, is_afk=0 
                WHERE user_id=?
            ''', (user_id,))
            conn.commit()
        
        needs.update_needs(user_id)  # Обновляем потребности
        
        return await func(update, context)
    return wrapper

@update_activity
async def start(update: Update, context: CallbackContext) -> int:
    user_id = update.message.from_user.id
    if user_exists(user_id):
        keyboard = [
            [InlineKeyboardButton("📊 Профиль", callback_data='show_profile')],
            [InlineKeyboardButton("🕒 Время", callback_data='show_time'),
             InlineKeyboardButton("🌤 Погода", callback_data='show_weather')]
        ]
        await update.message.reply_text("Главное меню:", 
            reply_markup=InlineKeyboardMarkup(keyboard))
        return ConversationHandler.END
        
    await update.message.reply_text(
        "🎮 Добро пожаловать!\n➖➖➖➖➖➖➖➖➖➖\n"
        "📛 Введите уникальный никнейм (2-12 символов):"
    )
    return GET_NICKNAME

@update_activity
async def process_nickname(update: Update, context: CallbackContext) -> int:
    user_id = update.message.from_user.id
    nickname = update.message.text.strip()

    if not 2 <= len(nickname) <= 12:
        await update.message.reply_text("❌ Никнейм должен содержать от 2 до 12 символов!")
        return GET_NICKNAME

    try:
        with sqlite3.connect('game.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute('INSERT INTO users (user_id, username) VALUES (?, ?)', 
                          (user_id, nickname))
            conn.commit()
            
            cursor.execute('SELECT player_id FROM users WHERE user_id = ?', (user_id,))
            player_id = cursor.fetchone()[0]
            await update.message.reply_text(
                f"✅ Регистрация успешна!\n🆔 ID: ID{player_id}\n👤 Ник: {nickname}\n"
                f"➖➖➖➖➖➖➖➖➖➖\nИспользуйте /profile для статистики"
            )
    except sqlite3.IntegrityError:
        await update.message.reply_text("⚠️ Этот никнейм уже занят! Выберите другой:")
        return GET_NICKNAME

    await show_profile(update, context)
    return ConversationHandler.END

@update_activity
async def button_handler(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer()
    if query.data == 'show_profile':
        await show_profile(update, context)
    elif query.data == 'show_time':
        await show_time(update, context)
    elif query.data == 'show_weather':
        await show_weather(update, context)

@update_activity
async def show_profile(update: Update, context: CallbackContext) -> None:
    user_id = update.callback_query.from_user.id if update.callback_query else update.message.from_user.id
    with sqlite3.connect('game.db', timeout=10) as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT player_id, username, level, experience, money, skills, location, is_afk 
            FROM users WHERE user_id = ?
        ''', (user_id,))
        result = cursor.fetchone()
    
    if not result:
        await context.bot.send_message(chat_id=user_id, text="❌ Вы не зарегистрированы!")
        return
        
    p_id, username, lvl, exp, money, skills, location, is_afk = result
    exp_needed = 4 * (lvl + 1)
    afk_status = "🔴 AFK" if is_afk else "🟢 Активен"

    needs_data = needs.get_needs(user_id)
    food = needs_data["food"] if needs_data else 0
    water = needs_data["water"] if needs_data else 0
    sleep = needs_data["sleep"] if needs_data else 0

    await context.bot.send_message(
        chat_id=user_id,
        text=f"{afk_status}\n📊 Профиль ID{p_id}\n👤 Ник: {username}\n"
             f"🏅 Уровень: {lvl}\n🔋 Опыт: {exp}/{exp_needed}\n💰 Деньги: {money}\n"
             f"🎯 Навыки: {skills}\n📍 Местоположение: {location}\n"
             f"🍖 Еда: {food}%\n💧 Вода: {water}%\n😴 Сон: {sleep}%"
    )

async def auto_income_task(context: CallbackContext):
    try:
        with sqlite3.connect('game.db') as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT user_id, skills, last_active, is_afk FROM users')
            
            now = datetime.now()
            for user_id, skills, last_active_str, is_afk in cursor.fetchall():
                last_active = datetime.strptime(last_active_str, "%Y-%m-%d %H:%M:%S")
                hours_inactive = (now - last_active).total_seconds() / 3600
                
                if hours_inactive >= 24:
                    if not is_afk:
                        cursor.execute('UPDATE users SET is_afk = 1 WHERE user_id = ?', (user_id,))
                        await context.bot.send_message(
                            chat_id=user_id,
                            text="🔴 Вы перешли в режим AFK из-за неактивности!"
                        )
                    continue
                
                cursor.execute('''
                    UPDATE users SET 
                        experience = experience + 1,
                        money = money + ?
                    WHERE user_id = ?''', 
                    (skills * 10, user_id))
                
                await add_experience(user_id, 1)
                
                try:
                    await context.bot.send_message(
                        chat_id=user_id,
                        text=f"🔄 Автоначисление: +{skills * 10} монет, +1 опыт"
                    )
                except Exception as e:
                    print(f"Ошибка отправки {user_id}: {e}")
            
            conn.commit()
    except Exception as e:
        print(f"Ошибка в auto_income_task: {e}")

@update_activity
async def export_stats(update: Update, context: CallbackContext) -> None:
    chat_id = update.message.chat_id
    
    if chat_id not in CHAT_ID_WHITELIST:
        await update.message.reply_text("❌ У вас нет прав на использование этой команды.")
        return

    if context.args:
        try:
            target_user_id = int(context.args[0])
            filename = get_export_filename()
            
            with sqlite3.connect('game.db') as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT player_id, username, level, experience, money, skills, location 
                    FROM users WHERE user_id = ?
                ''', (target_user_id,))
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Player"
                ws.append(["ID", "Никнейм", "Уровень", "Опыт", "Деньги", "Навыки", "Местоположение"])
                
                for row in cursor.fetchall():
                    ws.append(row)
                
                wb.save(filename)
            
            await update.message.reply_document(
                document=open(filename, "rb"),
                caption=f"📊 Статистика игрока {target_user_id} экспортирована в Excel"
            )
            return
            
        except (ValueError, IndexError):
            await update.message.reply_text("❌ Неверный формат ID. Используйте: /export_stats [ID]")
            return

    filename = get_export_filename()
    export_to_excel(filename)
    await update.message.reply_document(
        document=open(filename, "rb"),
        caption="📊 Статистика всех игроков экспортирована в Excel"
    )

@update_activity
async def import_stats(update: Update, context: CallbackContext) -> None:
    file = await context.bot.get_file(update.message.document.file_id)
    await file.download_to_drive("imported_stats.xlsx")
    
    try:
        import_from_excel("imported_stats.xlsx")
        await update.message.reply_text("✅ Данные успешно импортированы!")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка импорта: {str(e)}")

async def auto_export_stats(context: CallbackContext) -> None:
    filename = get_export_filename()
    export_to_excel(filename)
    print(f"✅ Автоэкспорт выполнен: {os.path.basename(filename)}")

@update_activity
async def set_location(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    if not user_exists(user_id):
        await update.message.reply_text("❌ Вы не зарегистрированы! Используйте /start.")
        return

    if not context.args:
        await update.message.reply_text(
            "❌ Укажите местоположение из списка.\n"
            "Пример: /set_location Лес\n"
            "Посмотреть доступные локации: /locations"
        )
        return

    location = " ".join(context.args).strip()
    
    if location not in ALLOWED_LOCATIONS:
        await update.message.reply_text(
            "❌ Неизвестная локация!\n"
            "Посмотрите доступные локации: /locations"
        )
        return

    with sqlite3.connect('game.db', timeout=10) as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE users SET location = ? WHERE user_id = ?', (location, user_id))
        conn.commit()

    await update.message.reply_text(
        f"✅ Вы переместились в: {location}\n"
        f"Используйте /profile для просмотра характеристик"
    )

@update_activity
async def show_location(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    if not user_exists(user_id):
        await update.message.reply_text("❌ Вы не зарегистрированы! Используйте /start.")
        return

    with sqlite3.connect('game.db', timeout=10) as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT location FROM users WHERE user_id = ?', (user_id,))
        location = cursor.fetchone()[0]

    await update.message.reply_text(f"📍 Ваше текущее местоположение: {location}")

@update_activity
async def show_locations(update: Update, context: CallbackContext) -> None:
    locations_list = "\n".join(f"📍 {loc}" for loc in ALLOWED_LOCATIONS)
    await update.message.reply_text(
        f"🗺 Доступные локации:\n{locations_list}\n\n"
        "Используйте /set_location [название] для перемещения"
    )

# ====================== КОМАНДЫ ПОТРЕБНОСТЕЙ =====================
@update_activity
async def eat(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    if not user_exists(user_id):
        await update.message.reply_text("❌ Вы не зарегистрированы!")
        return
    
    if needs.restore_need(user_id, "food", 20):
        await update.message.reply_text("🍖 Вы поели! +20% к еде.")
    else:
        await update.message.reply_text("❌ Ошибка при восстановлении еды.")

@update_activity
async def drink(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    if not user_exists(user_id):
        await update.message.reply_text("❌ Вы не зарегистрированы!")
        return
    
    if needs.restore_need(user_id, "water", 20):
        await update.message.reply_text("💧 Вы попили! +20% к воде.")
    else:
        await update.message.reply_text("❌ Ошибка при восстановлении воды.")

@update_activity
async def rest(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    if not user_exists(user_id):
        await update.message.reply_text("❌ Вы не зарегистрированы!")
        return
    
    if needs.restore_need(user_id, "sleep", 20):
        await update.message.reply_text("😴 Вы поспали! +20% к сну.")
    else:
        await update.message.reply_text("❌ Ошибка при восстановлении сна.")

# ====================== ВРЕМЯ И ПОГОДА ===========================
def get_moscow_time() -> str:
    now = datetime.utcnow() + timedelta(hours=MOSCOW_TIMEZONE_OFFSET)
    return now.strftime("%d.%m.%Y %H:%M")

async def get_moscow_weather() -> str:
    url = f"http://api.openweathermap.org/data/2.5/weather?q=Moscow&appid={OPENWEATHER_API_KEY}&units=metric&lang=ru"
    try:
        response = requests.get(url)
        data = response.json()
        if response.status_code == 200:
            weather = data["weather"][0]["description"].capitalize()
            temp = data["main"]["temp"]
            humidity = data["main"]["humidity"]
            wind_speed = data["wind"]["speed"]
            return (
                f"🌤 Погода в Москве:\n"
                f"• Состояние: {weather}\n"
                f"• Температура: {temp}°C\n"
                f"• Влажность: {humidity}%\n"
                f"• Ветер: {wind_speed} м/с"
            )
        else:
            return "❌ Не удалось получить данные о погоде"
    except Exception as e:
        print(f"Ошибка при получении погоды: {e}")
        return "❌ Ошибка подключения к сервису погоды"

@update_activity
async def show_time(update: Update, context: CallbackContext) -> None:
    time_str = get_moscow_time()
    await update.message.reply_text(f"🕒 Текущее время в Москве: {time_str}")

@update_activity
async def show_weather(update: Update, context: CallbackContext) -> None:
    weather_str = await get_moscow_weather()
    await update.message.reply_text(weather_str)

# ====================== ЗАПУСК ПРИЛОЖЕНИЯ ========================
def main():
    init_db()
    app = Application.builder().token(BOT_TOKEN).build()
    
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={GET_NICKNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, process_nickname)]},
        fallbacks=[]
    )
    
    app.add_handler(conv_handler)
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(CommandHandler("profile", show_profile))
    app.add_handler(CommandHandler("export_stats", export_stats))
    app.add_handler(MessageHandler(filters.Document.FileExtension("xlsx"), import_stats))
    app.add_handler(CommandHandler("set_location", set_location))
    app.add_handler(CommandHandler("show_location", show_location))
    app.add_handler(CommandHandler("locations", show_locations))
    app.add_handler(CommandHandler("time", show_time))
    app.add_handler(CommandHandler("weather", show_weather))
    app.add_handler(CommandHandler("eat", eat))
    app.add_handler(CommandHandler("drink", drink))
    app.add_handler(CommandHandler("rest", rest))
    
    app.job_queue.run_repeating(auto_income_task, interval=3600, first=10)
    app.job_queue.run_repeating(auto_export_stats, interval=300, first=10)
    app.job_queue.run_daily(auto_cleanup_task, time=time(3, 0, 0))
    
    print("✅ Бот запущен! Для выхода Ctrl+C")
    app.run_polling()

if __name__ == "__main__":
    main()